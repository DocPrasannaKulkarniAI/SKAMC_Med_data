import streamlit as st
import anthropic
import base64
import json
import io
import re
from PIL import Image
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─── LOAD CONFIG ─────────────────────────────────────────────────────────────
try:
    # Cloud deployment — reads from Streamlit Secrets
    API_KEY     = st.secrets["ANTHROPIC_API_KEY"]
    AI_MODEL    = st.secrets.get("AI_MODEL", "claude-haiku-4-5-20251001")
    INSTITUTION = st.secrets.get("INSTITUTION", "Sri Kalabyraveshwara Swamy Ayurvedic Medical College, Bangalore")
except Exception:
    try:
        # Local fallback — reads from config.py
        import config
        API_KEY     = config.ANTHROPIC_API_KEY
        AI_MODEL    = config.AI_MODEL
        INSTITUTION = config.INSTITUTION
    except ImportError:
        st.error("❌ config.py not found. Please set up your API key.")
        st.stop()

if "PASTE-YOUR-KEY-HERE" in API_KEY or not API_KEY.startswith("sk-ant-"):
    st.error("❌ Please open config.py and paste your Anthropic API key into ANTHROPIC_API_KEY.")
    st.stop()

# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Ayurveda OPD Extractor",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── CUSTOM CSS ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .stApp { background-color: #f5f0e8; }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a3a0f 0%, #2d5016 100%);
    }
    [data-testid="stSidebar"] * { color: #d4f0a0 !important; }
    [data-testid="stSidebar"] label { color: #a8c97f !important; }

    .opd-header {
        background: linear-gradient(135deg, #1a3a0f, #2d5016, #4a7c20);
        border-radius: 12px;
        padding: 20px 28px;
        margin-bottom: 20px;
        border-left: 6px solid #d4af37;
    }
    .opd-header h1 { color: #fff; margin: 0; font-size: 24px; }
    .opd-header p  { color: #a8c97f; margin: 4px 0 0; font-size: 13px; }

    .status-success {
        background: #e8f5e9; border-left: 4px solid #2d5016;
        padding: 10px 14px; border-radius: 6px; margin: 8px 0; font-size: 13px; color: #1a3a0f;
    }
    .status-error {
        background: #fde8e8; border-left: 4px solid #c0392b;
        padding: 10px 14px; border-radius: 6px; margin: 8px 0; font-size: 13px; color: #922b21;
    }
    .status-info {
        background: #e8f0fe; border-left: 4px solid #1a73e8;
        padding: 10px 14px; border-radius: 6px; margin: 8px 0; font-size: 13px; color: #174ea6;
    }
    .section-header {
        background: #2d5016; color: #d4af37;
        padding: 8px 16px; border-radius: 6px;
        font-size: 14px; font-weight: bold; margin: 16px 0 10px;
    }
    .stButton > button {
        background: linear-gradient(135deg, #2d5016, #4a7c20) !important;
        color: white !important; border: none !important;
        border-radius: 8px !important; font-weight: bold !important;
    }
    .stButton > button:hover {
        background: linear-gradient(135deg, #1a3a0f, #2d5016) !important;
        box-shadow: 0 4px 12px rgba(45,80,22,0.3) !important;
    }
    .footer {
        text-align: center; padding: 16px; color: #8b6914;
        font-size: 11px; margin-top: 32px; border-top: 1px solid #e8dfc0;
    }
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTS ────────────────────────────────────────────────────────────────
COLUMNS = [
    ("Sl_No",              "Sl.No"),
    ("Patient_ID",         "Patient_ID"),
    ("Registration_Year",  "Registration_Year"),
    ("Age",                "Age"),
    ("Gender",             "Gender"),
    ("District",           "District"),
    ("Occupation",         "Occupation"),
    ("Prakriti",           "Prakriti"),
    ("Lifestyle_Risk",     "Lifestyle_Risk"),
    ("Visit_ID",           "Visit_ID"),
    ("Visit_Date",         "Visit_Date"),
    ("Visit_Type",         "Visit_Type"),
    ("Department",         "Department"),
    ("Physician_ID",       "Physician_ID"),
    ("Visit_Year",         "Visit_Year"),
    ("Diagnosis_ID",       "Diagnosis_ID"),
    ("Ayurveda_Diagnosis", "Ayurveda_Diagnosis"),
    ("Disease_Category",   "Disease_Category"),
    ("Dosha_Dominance",    "Dosha_Dominance"),
    ("Disease_Duration",   "Disease_Duration"),
    ("Severity",           "Severity"),
    ("Observation_ID",     "Observation_ID"),
    ("Nadi",               "Nadi"),
    ("Jihva",              "Jihva"),
    ("Agni",               "Agni"),
    ("Mala",               "Mala"),
    ("Mutra",              "Mutra"),
    ("Sleep",              "Sleep"),
    ("Dosha",              "Dosha"),
    ("Dushya",             "Dushya"),
    ("Bala",               "Bala"),
    ("Kala",               "Kala"),
    ("Prakruti",           "Prakruti"),
    ("Satwa",              "Satwa"),
    ("Satmya",             "Satmya"),
    ("Vyasana",            "Vyasana"),
    ("Shabda",             "Shabda"),
    ("Sparsha",            "Sparsha"),
    ("Drik",               "Drik"),
    ("Akruti",             "Akruti"),
    ("Weight",             "Weight"),
    ("BP",                 "B.P"),
    ("Other_Investigation","Other_Investigation"),
]
COLUMN_KEYS   = [c[0] for c in COLUMNS]
COLUMN_LABELS = [c[1] for c in COLUMNS]

SYSTEM_PROMPT = """You are an expert Ayurvedic clinical data extraction specialist with deep knowledge of:
- Classical Ayurvedic terminology (Sanskrit, Devanagari script, and transliterated forms)
- Ayurvedic OPD/IPD registration forms used in Indian medical colleges and hospitals
- Both printed and handwritten medical records in mixed English-Sanskrit-Devanagari formats

CRITICAL INSTRUCTIONS:
1. Read EVERY handwritten entry carefully. Use Ayurvedic domain knowledge to infer ambiguous handwriting.
2. Devanagari/Hindi → transliterate to English. E.g. "मध्यम" → "Madhyama", "कफ-वात" → "Kapha-Vata".
3. Blank/dashed/"NIL" fields → return null.
4. Nadi: extract BPM and/or type (e.g. "72 bpm", "Vataja").
5. Mala/Mutra: frequency or quality (e.g. "1+/d", "5-6 times/day").
6. Visit_Type: OPD No "N" prefix → "New"; "O" prefix → "Old/Follow-up".
7. District: city/district extracted from Address.
8. Disease_Category: infer — one of: Musculoskeletal, Digestive, Metabolic, Neurological, Respiratory, Dermatological, Gynecological, Geriatric, Psychiatric, General.
9. Dosha_Dominance: primary dosha from Dosha field.
10. Lifestyle_Risk: from Vyasana; null if NIL.
11. Sleep/Severity/Disease_Duration: extract if mentioned; otherwise null.
12. Observation_ID = "OBS-" + Patient_ID. Diagnosis_ID = "DX-" + Patient_ID. Visit_ID = "V-" + Patient_ID.

Return ONLY valid JSON — no preamble, no markdown, no explanation:
{"Sl_No":1,"Patient_ID":"","Registration_Year":"","Age":"","Gender":"","District":"","Occupation":"","Prakriti":"","Lifestyle_Risk":null,"Visit_ID":"","Visit_Date":"","Visit_Type":"","Department":"","Physician_ID":"","Visit_Year":"","Diagnosis_ID":"","Ayurveda_Diagnosis":"","Disease_Category":"","Dosha_Dominance":"","Disease_Duration":null,"Severity":null,"Observation_ID":"","Nadi":"","Jihva":"","Agni":"","Mala":"","Mutra":"","Sleep":null,"Dosha":"","Dushya":"","Bala":"","Kala":"","Prakruti":"","Satwa":"","Satmya":"","Vyasana":null,"Shabda":"","Sparsha":"","Drik":"","Akruti":"","Weight":null,"BP":"","Other_Investigation":null}"""

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def image_to_base64(uploaded_file):
    data = uploaded_file.read()
    uploaded_file.seek(0)
    b64 = base64.standard_b64encode(data).decode("utf-8")
    ext = uploaded_file.name.split(".")[-1].lower()
    mt  = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png","webp":"image/webp"}
    return b64, mt.get(ext, "image/jpeg")


def extract_record(client, b64, media_type, sl_no):
    resp = client.messages.create(
        model=AI_MODEL,
        max_tokens=1024,
        system=SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": [
                {"type":"image","source":{"type":"base64","media_type":media_type,"data":b64}},
                {"type":"text","text":f"Extract all fields. Set Sl_No to {sl_no}. Return ONLY valid JSON."}
            ]
        }]
    )
    raw   = resp.content[0].text.strip()
    clean = re.sub(r"```(?:json)?|```", "", raw).strip()
    return json.loads(clean)


def build_excel(records, institution):
    wb = Workbook()
    ws = wb.active
    ws.title = "OPD_Data"

    hdr_fill = PatternFill("solid", fgColor="2D5016")
    alt_fill = PatternFill("solid", fgColor="F5F0E8")
    hdr_font = Font(bold=True, color="D4AF37", size=10)
    body_font= Font(size=10)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="C9A84C")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 36
    for ci, label in enumerate(COLUMN_LABELS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = center; c.border = border

    for ri, rec in enumerate(records, 2):
        ws.row_dimensions[ri].height = 22
        for ci, key in enumerate(COLUMN_KEYS, 1):
            val = rec.get(key, "")
            c   = ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")
            c.font = body_font
            c.alignment = center if ci <= 3 else left
            c.border = border
            if ri % 2 == 0:
                c.fill = alt_fill

    col_widths = {1:6,2:12,3:18,4:6,5:8,6:14,7:14,8:16,9:14,10:12,
                  11:14,12:12,13:12,14:20,15:10}
    for ci in range(1, len(COLUMN_KEYS)+1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(ci, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    meta = wb.create_sheet("Info")
    meta["A1"] = "Ayurveda OPD Data Extractor"
    meta["A2"] = f"Institution: {institution}"
    meta["A3"] = f"Generated: {datetime.now().strftime('%d-%b-%Y  %H:%M')}"
    meta["A4"] = f"Total Records: {len(records)}"
    meta["A5"] = f"Model: {AI_MODEL}"
    meta["A1"].font = Font(bold=True, size=13, color="2D5016")
    meta.column_dimensions["A"].width = 60

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


def to_dataframe(records):
    return pd.DataFrame([{lbl: r.get(k,"") for k,lbl in COLUMNS} for r in records])


# ─── SESSION STATE ────────────────────────────────────────────────────────────
if "records"         not in st.session_state: st.session_state.records = []
if "processed_names" not in st.session_state: st.session_state.processed_names = set()

client = anthropic.Anthropic(api_key=API_KEY)

# ─── SIDEBAR ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 🌿 SKAMC")
    st.markdown("**OPD Card Data Extractor**")
    st.divider()
    st.markdown(f"""
    **⚙️ Configuration**
    - Model : `{AI_MODEL.split('-')[1].capitalize()}`
    - API   : ✅ Connected
    """)
    st.divider()
    st.markdown("**📊 Session**")
    st.metric("Records Extracted", len(st.session_state.records))
    st.metric("Total Columns", len(COLUMNS))
    st.divider()
    if st.session_state.records:
        if st.button("🗑️ Clear All Records", use_container_width=True):
            st.session_state.records = []
            st.session_state.processed_names = set()
            st.rerun()
    st.divider()
    st.markdown("""
    **📋 Fields Captured**
    Demographics · Prakruti
    Nadi · Jihva · Agni
    Dosha · Dushya · Bala
    Diagnosis · Disease Category
    Vitals · Ashtasthana Pareeksha
    """)
    st.divider()
    st.markdown(f"""
    <div style='font-size:10px; color:#6db33f; text-align:center;'>
    {INSTITUTION}<br><br>
    Handles: Printed + Handwritten<br>Devanagari · Sanskrit · English
    </div>""", unsafe_allow_html=True)

# ─── HEADER ──────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="opd-header">
    <h1>🌿 Ayurveda OPD Card — AI Data Extractor</h1>
    <p>{INSTITUTION} &nbsp;|&nbsp; SKAMC</p>
    <p>Upload OPD card photos → AI reads printed + handwritten fields → Download Excel / CSV / JSON</p>
</div>""", unsafe_allow_html=True)

# ─── UPLOAD ──────────────────────────────────────────────────────────────────
st.markdown('<div class="section-header">📤 Upload OPD Card Images</div>', unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "Drag & drop OPD card images — JPG, PNG, WEBP — multiple files supported",
    type=["jpg","jpeg","png","webp"],
    accept_multiple_files=True,
    help="Each card image becomes one row in the output dataset."
)

# ─── PROCESS ─────────────────────────────────────────────────────────────────
if uploaded_files:
    new_files = [f for f in uploaded_files if f.name not in st.session_state.processed_names]

    if new_files:
        c1, c2 = st.columns([1,3])
        with c1:
            go = st.button(f"⚙️ Extract {len(new_files)} Card(s)", use_container_width=True)
        with c2:
            st.markdown(f'<div class="status-info">📋 {len(new_files)} new image(s) ready</div>',
                        unsafe_allow_html=True)
        if go:
            bar    = st.progress(0)
            status = st.empty()
            for i, f in enumerate(new_files):
                status.markdown(
                    f'<div class="status-info">⚙️ Reading <strong>{f.name}</strong>'
                    f' ({i+1}/{len(new_files)})…</div>', unsafe_allow_html=True)
                try:
                    b64, mt = image_to_base64(f)
                    rec     = extract_record(client, b64, mt, len(st.session_state.records)+1)
                    st.session_state.records.append(rec)
                    st.session_state.processed_names.add(f.name)
                    status.markdown(
                        f'<div class="status-success">✅ <strong>{f.name}</strong> &nbsp;|&nbsp;'
                        f' Patient: {rec.get("Patient_ID","?")} &nbsp;|&nbsp;'
                        f' Diagnosis: {rec.get("Ayurveda_Diagnosis","?")}</div>',
                        unsafe_allow_html=True)
                except json.JSONDecodeError as e:
                    st.session_state.records.append({"Sl_No":len(st.session_state.records)+1,
                                                      "Patient_ID":f"PARSE_ERROR:{f.name}"})
                    st.markdown(f'<div class="status-error">❌ JSON error — {f.name}: {e}</div>',
                                unsafe_allow_html=True)
                except Exception as e:
                    st.session_state.records.append({"Sl_No":len(st.session_state.records)+1,
                                                      "Patient_ID":f"ERROR:{f.name}"})
                    st.markdown(f'<div class="status-error">❌ Failed — {f.name}: {e}</div>',
                                unsafe_allow_html=True)
                bar.progress((i+1)/len(new_files))
            status.markdown(
                f'<div class="status-success">🎉 All done! '
                f'<strong>{len(new_files)}</strong> card(s) processed. '
                f'Total: <strong>{len(st.session_state.records)}</strong> records</div>',
                unsafe_allow_html=True)
            st.rerun()
    else:
        st.markdown('<div class="status-info">ℹ️ All uploaded files already processed this session.</div>',
                    unsafe_allow_html=True)

    with st.expander(f"🖼️ Image Previews ({len(uploaded_files)})", expanded=False):
        cols = st.columns(min(len(uploaded_files), 4))
        for i, f in enumerate(uploaded_files):
            with cols[i % 4]:
                img = Image.open(f); f.seek(0)
                st.image(img, caption=f.name, use_container_width=True)

# ─── RESULTS ─────────────────────────────────────────────────────────────────
if st.session_state.records:
    st.markdown(f'<div class="section-header">📊 Extracted Records — {len(st.session_state.records)} card(s)</div>',
                unsafe_allow_html=True)
    df = to_dataframe(st.session_state.records)
    st.dataframe(df, use_container_width=True, height=300)

    with st.expander("📈 Quick Summary Stats", expanded=False):
        c1,c2,c3,c4 = st.columns(4)
        with c1:
            st.markdown("**Gender**")
            st.dataframe(df["Gender"].value_counts(), use_container_width=True)
        with c2:
            st.markdown("**Department**")
            st.dataframe(df["Department"].value_counts(), use_container_width=True)
        with c3:
            st.markdown("**Disease Category**")
            st.dataframe(df["Disease_Category"].value_counts(), use_container_width=True)
        with c4:
            st.markdown("**Dosha Dominance**")
            st.dataframe(df["Dosha_Dominance"].value_counts(), use_container_width=True)

    st.markdown('<div class="section-header">📥 Download Extracted Data</div>', unsafe_allow_html=True)
    d1,d2,d3 = st.columns(3)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    with d1:
        st.download_button("📥 Excel (.xlsx)",
            data=build_excel(st.session_state.records, INSTITUTION),
            file_name=f"OPD_Records_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with d2:
        st.download_button("📄 CSV",
            data=df.to_csv(index=False).encode("utf-8"),
            file_name=f"OPD_Records_{ts}.csv",
            mime="text/csv", use_container_width=True)
    with d3:
        st.download_button("🗂️ JSON",
            data=json.dumps(st.session_state.records, indent=2, ensure_ascii=False).encode("utf-8"),
            file_name=f"OPD_Records_{ts}.json",
            mime="application/json", use_container_width=True)

# ─── EMPTY STATE ─────────────────────────────────────────────────────────────
if not st.session_state.records and not uploaded_files:
    st.markdown("""
    <div style='text-align:center; padding:48px; background:white; border-radius:12px;
                border:2px dashed #c9a84c; margin:16px 0;'>
        <div style='font-size:56px; margin-bottom:12px;'>📋</div>
        <div style='font-size:18px; font-weight:bold; color:#2d5016;'>Ready to Extract</div>
        <div style='color:#7a6030; margin-top:8px; font-size:13px; line-height:2.2;'>
            1. Upload one or more OPD card photos above<br>
            2. Click <strong>Extract Cards</strong><br>
            3. Review the table — then download Excel / CSV / JSON
        </div>
    </div>""", unsafe_allow_html=True)

# ─── FOOTER ──────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="footer">
    🌿 SKAMC &nbsp;|&nbsp; {INSTITUTION} &nbsp;|&nbsp;
    Powered by Claude AI (Anthropic) &nbsp;|&nbsp; <em>For research & academic use</em>
</div>""", unsafe_allow_html=True)
